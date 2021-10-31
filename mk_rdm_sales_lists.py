# conding:utf-8
"""
可以实现的功能：
根据金额随机生产订单列表，订单列表总金额精确地与目标金额一致
按一定概率分布随机生产商品种类，按一定概率分布随机生产每种商品购买的个数
需要读取E:\\1D\\data1\\目录下 日文商品名-顾客姓名源文件.xlsx中的文件，随机选取商品名和顾客姓名
类MakeRdmLists需要3个参数：int金额money_goal，str日期date，要保存的int文件名类型type_type，为1时是支付宝，为2时是淘宝订单，为其他值时为微信
"""
import openpyxl
import random
import numpy


class MakeRdmLists(object):
    def __init__(self, money_goal, date, type_type):  # 需要参数int金额money_goal，str日期date，要保存的int文件名类型type_type
        self.money_goal = money_goal
        self.date = date
        if type_type == 1:
            self.title = 'Alipay-' + date + '.xlsx'
        elif type_type == 2:
            self.title = 'Taobao-' + date + '.xlsx'
        else:
            self.title = 'Wechat-' + date + '.xlsx'
        self.xl = openpyxl.Workbook()  # 新建excel
        self.nsht = self.xl[self.xl.active.title]  # 新建excel的sheet

    # 随机选一个顾客，随机确定商品种类数，随机确定每种商品的购买个数，写入顾客信息，购买产品名、数量、价格、合计，统计订单总金额
    def mk_rdm_order(self):
        xl_res = openpyxl.load_workbook('E:\\1D\\data1\\日文商品名-顾客姓名源文件.xlsx')  # 打开源文件
        sht_res = xl_res[xl_res.active.title]  # 源文件的sheet
        max_d = max((dd.row for dd in sht_res['D'] if dd.value))  # 找到源中D列最大行，即顾客名最大行
        max_b = max((dd.row for dd in sht_res['B'] if dd.value))  # 找到源中B列最大行，即商品名最大行
        row_n = 2  # 新建excel里的行数-2遍赋值是防止：while判定失败时，row_n，num_lists会因为没有赋值而警告
        num_lists = 1  # 订单号-2遍赋值是防止：while判定失败时，row_n，num_lists会因为没有赋值而警告
        # money_lists = 0  # 随机订单的总金额-2遍赋值是防止：while判定失败时，row_n，num_lists会因为没有赋值而警告
        diff = -1  # 目标金额与总金额的差值，定-1是为了第一次时进入下面的while
        while diff < 0:  # 如果第二层while后diff为负值，说明总额超过了目标金额，要重做，为正值不重做
            row_n = 2  # 新建excel里的行数
            num_lists = 1  # 订单号
            money_lists = 0  # 随机订单的总金额
            diff = -1  # 重新赋值-1是为了：重做时能够进入第二层while
            self.xl = openpyxl.Workbook()  # 需要新建excel，不然excel里会有重做前一次里写入的数据
            self.nsht = self.xl[self.xl.active.title]  # 新建excel的sheet
            self.nsht['A1'] = '番号'  # 新建excel
            self.nsht['B1'] = '販売日'
            self.nsht['C1'] = '商品'
            self.nsht['D1'] = '数量'
            self.nsht['E1'] = '金額'
            self.nsht['F1'] = '販売先'
            self.nsht['G1'] = '住所'
            self.nsht['H1'] = '数量*金額'
            # 第二层while创建所有的订单，直到diff小于500了。每次循环一次，新建一个订单。
            while diff > 500 or diff == -1:  # diff==-1是为了第一次或重做时能够进入
                random.seed()  # 初始化随机种子
                guest_row = random.randint(2, max_d)  # 随机找一个顾客的行数
                p2 = numpy.array([0.4, 0.2, 0.2, 0.15, 0.05])  # 产品种类数的分布概率
                good_sort = numpy.random.choice([1, 2, 3, 4, 5], p=p2.ravel())  # 随机定一个顾客购买的产品种类
                self.nsht[f'A{str(row_n)}'] = num_lists  # 写入订单编号
                for i in range(good_sort):  # 每个i为一种商品，对应新建excel里一行。整个for为一个顾客的订单
                    random.seed()
                    p1 = numpy.array([0.85, 0.1, 0.05])  # 按一定概率选择随机数，用来随机定商品数量
                    good_row = random.randint(2, max_b)  # 从源中随机找一个商品的行数，randint包括起止点，所以不能+1
                    self.nsht[f'C{str(row_n)}'] = sht_res[f'B{str(good_row)}'].value  # 写入商品名
                    good_num = numpy.random.choice([1, 2, 3], p=p1.ravel())  # 随机确定一个商品的购买数量
                    self.nsht[f'D{str(row_n)}'] = good_num  # 写入商品数量
                    self.nsht[f'E{str(row_n)}'] = round(float(sht_res[f'C{str(good_row)}'].value), 2)  # 写入价格
                    self.nsht[f'H{str(row_n)}'] = good_num * self.nsht[f'E{str(row_n)}'].value  # 写入数量*金額
                    self.nsht[f'F{str(row_n)}'] = sht_res[f'D{str(guest_row)}'].value  # 写入随机到的顾客姓名
                    self.nsht[f'G{str(row_n)}'] = sht_res[f'E{str(guest_row)}'].value  # 写入顾客住址
                    money_lists = round(money_lists + self.nsht[f'H{str(row_n)}'].value, 2)  # 将本行合计金额加入总金额
                    row_n += 1  # 行下移1
                diff = round(self.money_goal - money_lists, 2)  # diff赋值真正含义，为目标金额与实际总金额的差，每个while验证一次
                num_lists += 1  # 订单号+1
                if diff < 0:  # 如果上一个订单时diff还>500，这一次就变负值了，执行
                    print(f'diff为{diff}，重新生成数据')
                    break  # 直接跳出第二层while，会到第一层while验证

        print(f"最终diff为：{diff}")
        diff_id_rdm = random.randint(2, 13)  # 随机生成一个小额商品的行数
        self.nsht[f'F{str(row_n)}'] = sht_res[f'D{str(diff_id_rdm)}'].value  # 写入随机到的顾客姓名
        self.nsht[f'G{str(row_n)}'] = sht_res[f'E{str(diff_id_rdm)}'].value  # 写入顾客住址
        self.nsht[f'D{str(row_n)}'] = 1  # 写入商品数量1
        self.nsht[f'H{str(row_n)}'] = diff  # 写入数量*金額
        self.nsht[f'E{str(row_n)}'] = diff  # 写入差值作为价格
        self.nsht[f'A{str(row_n)}'] = num_lists  # 写入订单编号
        if diff < 100:  # 如果差值小于100，用小小额补充商品
            self.nsht[f'C{str(row_n)}'] = sht_res[f'G{str(diff_id_rdm)}'].value  # 写入小小额补充商品名
        else:  # 如果差值大于100小于500，用小大额补充商品
            self.nsht[f'C{str(row_n)}'] = sht_res[f'F{str(diff_id_rdm)}'].value  # 写入小大额补充商品名
        print(f'订单数为：{num_lists}，最后行为：{row_n}, 填充订单金额为：{diff}')
        self.xl.save(self.title)
        return num_lists  # num_lists即为订单数，后面根据订单数生成日期

    def add_date(self, num_all):  # 根据订单数生成日期
        list_date = list(self.date)  # 将日期拆解为单个字母的列表
        month = list_date[4] + list_date[5]  # 获得月份
        year = list_date[0] + list_date[1] + list_date[2] + list_date[3]  # 获得年
        day_in_month = 31
        if month == '02':
            day_in_month = 28  # 月数不同，一个月的天数不同
        # 重大错误：如果写成：if month == '04' or '06' or '09' or '11':，则2月份的时候，day_in_month会被改为30
        elif month == '04' or month == '06' or month == '09' or month == '11':
            day_in_month = 30

        day_rdm_list = []  # 用来存放随机生成的num个订单对应的日期
        for _ in range(num_all):  # 根据订单数生成天数的字符串，每个_生成一个日期str
            random.seed()
            day_rdm = random.randint(1, day_in_month)  # 随机生成天数
            if day_rdm < 10:  # 小于10，前面加0
                day_rdm_list.append('0' + str(day_rdm))
            else:
                day_rdm_list.append(str(day_rdm))
        day_rdm_list.sort()  # 从小到大排序

        num = 1  # 加日期时的订单编号
        row = 2  # 行号
        while num <= num_all:  # 每次循环一次，生成一个订单的日期,num是处理好的订单数
            random.seed()
            hour_rdm = random.randint(9, 23)  # 生成随机的hour整数
            minute_rdm = random.randint(0, 60)  # 生成随机的minute整数
            second_rdm = random.randint(0, 60)  # 生成随机的second整数
            if hour_rdm < 10:
                hour_rdm = '0' + str(hour_rdm)  # 小于10，前面加0
            else:
                hour_rdm = str(hour_rdm)  # 生成随机的hour字符串
            if minute_rdm < 10:
                minute_rdm = '0' + str(minute_rdm)  # 小于10，前面加0
            else:
                minute_rdm = str(minute_rdm)  # 生成随机的minute字符串
            if second_rdm < 10:
                second_rdm = '0' + str(second_rdm)  # 小于10，前面加0
            else:
                second_rdm = str(second_rdm)  # 生成随机的second字符串

            if self.nsht[f'A{str(row)}'].value:  # 如果订单编号不为空，则写入生成的日期str
                self.nsht[f'B{str(row)}'] = year + '-' + month + '-' + day_rdm_list[num - 1] + \
                                                  " " + hour_rdm + ':' + minute_rdm + ':' + second_rdm
                num += 1  # 订单编号非空时+1
            row += 1  # 行号+1
        self.xl.save(self.title)

    def run(self):
        num_lists = self.mk_rdm_order()
        self.add_date(num_lists)


if __name__ == '__main__':
    date1 = '202002'  # 在这里修改 要保存的文件名称里的日期
    goal = 45678  # 目标金额，在这里修改 这个月的销售额
    type1 = 1  # 为1时是支付宝，为2时是淘宝订单，为其他值时为微信

    mkrdmlists = MakeRdmLists(goal, date1, type1)
    mkrdmlists.run()
