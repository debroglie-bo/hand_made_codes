#  conding:utf-8
"""
读取导出的东邦详情xlsx文件，统计其中的相同条形码的数量，品名
读取管家婆导出的‘商品列表.xlsx’文件，读取品名，并根据品名利用模块Discount获取折扣
可以读取 E:\\圆通物流东邦\\文件夹下的文件如：转运包裹详情表_20211030.xlsx
保存文件到 E:\\1D\\data1\\输出金额_20211030.xlsx
商品列表.xlsx简称为源
"""

import openpyxl
from Discount import discount  # 自定义的模块


class Export(object):
    def __init__(self):
        # self.date = input("请输入批次月日日期：\n")
        self.date = '0929'
        self.file_name = "E:\\圆通物流东邦\\转运包裹详情表_2021" + str(self.date) + ".xlsx"
        self.nxl = openpyxl.Workbook()  # 新建excel文件，文件作为一个属性，方便后面保存excel文件
        self.nsht = self.nxl[self.nxl.active.title]  # 选取活动sheet，openpyxl文件作为属性，方便后面写入excel

    def read_cal(self):
        sht = openpyxl.load_workbook(self.file_name)['Worksheet']  # 东邦详情excel
        sht_s = openpyxl.load_workbook('E:\\1D\\data1\\商品列表.xlsx')['商品']  # 打开商品列表excel
        max_row_s = sht_s.max_row  # 源的最大行数
        i = 2  # i是东邦excel里的行
        num_list_all = []  # 用来记录东邦详情excel里的商品条形码，有重复
        while sht[f'D{i}'].value:
            # num_list_all.append(str(int(sht[f'D{i}'].value)).zfill(13))  # 将excel格子的值去掉前后的引号，先转为int，再转str，补齐13位
            num_list_all.append(''.join([n for n in sht[f'D{i}'].value if n.isdigit()]))  # 只提取数字，这种方法有普遍性
            i += 1  # 循环结束后i=东邦excel的最大行+1
        num_list = list(set(num_list_all))  # 用集合来去重，再转为list
        name = []  # 用来存放商品名
        price_disc = {}
        all_price = 0
        for j in range(len(num_list)):  # j代表新excel中的行数，每个j代表一种商品
            num = 0  # num是统计的某种商品数量
            price_disc[j] = [0.0, 1.0]  # price_disc是字典，键是条形码序号，值是二维列表，float[0]放折扣, float[1]放管家婆价格
            name.append('0')  # 每一个j，name增加一个元素，以便后面修改对应的元素为从东邦详情excel里读取的商品名
            # num = (num + int(sht[f'F{m}'].value) for m in range(1, i) if num_list[j] ==
            # ''.join([n for n in sht[f'D{m}'].value if n.isdigit()]))
            for m in range(2, i):  # 遍历东邦详情中的行
                if num_list[j] == ''.join([n for n in sht[f'D{m}'].value if n.isdigit()]):
                    name[j] = sht[f'E{m}'].value  # 保存商品名
                    num = num + int(sht[f'F{m}'].value)  # 计算商品数量
            self.nsht[f'D{j + 2}'] = str(num)  # 写入商品总数量
            self.nsht[f'B{j + 2}'] = str(num_list[j])  # 写入条形码
            self.nsht[f'C{j + 2}'] = name[j]  # 写入东邦商品名

            for s in range(2, max_row_s + 1):  # 遍历源里查找条形码，每个s为一个行
                if num_list[j] == ''.join([n for n in sht_s[f'E{s}'].value if n.isdigit()]):
                    # num = num + int(sht[f'F{s}'].value)
                    self.nsht[f'G{j + 2}'] = sht_s[f'O{s}'].value  # 写入品名/管家婆
                    # 根据管家婆商品名写入折扣price_disc[j][0]，
                    # 按理说[0]处应该有警报，因为本身类型为int，而这里赋值的float，这里没有警报是因为不知道discount返回个什么东西
                    price_disc[j][0] = discount(sht_s[f'O{s}'].value)
                    self.nsht[f'H{j + 2}'] = price_disc[j][0]  # 根据管家婆商品名写入折扣
                    # 写入价格/管家婆，保留2位price_disc[j][1]
                    price_disc[j][1] = round(float(sht_s[f'U{s}'].value) * 1.1, 2)
                    self.nsht[f'F{j + 2}'] = price_disc[j][1]  # 写入价格/管家婆
                    break  # 源里条形码不重复，找到就可以终止循环了
            if price_disc[j][0] and price_disc[j][1]:  # 避免空值时报错
                all_price = round(all_price + price_disc[j][0] * price_disc[j][1] * int(num), 1)
        print(f"总金额为：{all_price}")
        self.nsht[f'I{len(num_list) + 3}'] = all_price
        self.nsht[f'B{len(num_list) + 3}'] = '合计'

    def run(self):
        self.nsht['A1'] = '编号'
        self.nsht['B1'] = '商品条形码'
        self.nsht['C1'] = '品名/东邦'
        self.nsht['D1'] = '数量'
        self.nsht['E1'] = '价格/东邦'
        self.nsht['F1'] = '价格/管家婆'
        self.nsht['G1'] = '品名/管家婆'
        self.nsht['H1'] = '折扣'
        self.nsht['I1'] = '合计'
        self.read_cal()
        self.nxl.save("E:\\1D\\data1\\输出金额_2021" + str(self.date) + ".xlsx")  # 保存excel文件


if __name__ == '__main__':
    export = Export()
    export.run()
    print('Done!')
